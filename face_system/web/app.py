"""
Flask Web Server + Excel Export
"""
from flask import Flask, jsonify, request, send_file, make_response
from datetime import datetime, date, timedelta
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .templates import get_dashboard_html

def create_web_app(face_system):
    """Flask app yaratish"""
    app = Flask(__name__)
    
    @app.route('/')
    def dashboard():
        """Asosiy dashboard"""
        return get_dashboard_html()
    
    @app.route('/api/dashboard-data')
    def api_dashboard_data():
        """Dashboard uchun barcha ma'lumotlar"""
        try:
            # Asosiy statistika
            status_data = {
                'is_running': face_system.is_running,
                'current_users': list(face_system.current_users.keys()),
                'total_users': len(face_system.known_face_names),
                'current_time': datetime.now().isoformat()
            }
            
            # Foydalanuvchilar statistikasi
            users_stats = face_system.get_user_stats()
            
            # Bugungi hisobot
            daily_summary = face_system.data_manager.get_daily_summary(date.today())
            
            # Bugungi kirishlar soni
            today_entries = 0
            if daily_summary.get('users'):
                for user_data in daily_summary['users'].values():
                    today_entries += len(user_data.get('sessions', []))
            
            return jsonify({
                'status': status_data,
                'users': users_stats,
                'daily_summary': daily_summary,
                'today_entries': today_entries
            })
        except Exception as e:
            return jsonify({'error': str(e)})
    
    @app.route('/api/status')
    def api_status():
        """Tizim holati"""
        return jsonify({
            'is_running': face_system.is_running,
            'current_users': list(face_system.current_users.keys()),
            'total_users': len(face_system.known_face_names),
            'current_time': datetime.now().isoformat()
        })
    
    @app.route('/api/users')
    def api_users():
        """Foydalanuvchilar statistikasi"""
        try:
            return jsonify(face_system.get_user_stats())
        except Exception as e:
            return jsonify({'error': str(e)})
    
    @app.route('/api/daily-summary')
    def api_daily_summary():
        """Kunlik hisobot"""
        try:
            date_param = request.args.get('date')
            if date_param:
                target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            else:
                target_date = date.today()
            
            summary = face_system.data_manager.get_daily_summary(target_date)
            return jsonify(summary)
        except Exception as e:
            return jsonify({'error': str(e)})
    
    @app.route('/api/weekly-summary')
    def api_weekly_summary():
        """Haftalik hisobot"""
        try:
            weekly_data = []
            for i in range(7):
                target_date = date.today() - timedelta(days=i)
                daily_summary = face_system.data_manager.get_daily_summary(target_date)
                weekly_data.append(daily_summary)
            
            return jsonify(weekly_data)
        except Exception as e:
            return jsonify({'error': str(e)})
    
    @app.route('/api/logs')
    def api_logs():
        """Tizim loglari"""
        try:
            recent_logs = face_system.data_manager.logs_data[-50:] if face_system.data_manager.logs_data else []
            return jsonify(recent_logs)
        except Exception as e:
            return jsonify({'error': str(e)})
    
    @app.route('/api/backup', methods=['POST'])
    def api_backup():
        """Backup yaratish"""
        backup_path = face_system.data_manager.create_backup()
        if backup_path:
            return jsonify({'success': True, 'backup_path': backup_path})
        else:
            return jsonify({'success': False, 'error': 'Backup yaratishda xatolik'})
    
    @app.route('/api/export-excel')
    def api_export_excel():
        """Kunlik ma'lumotlarni Excel formatda yuklab olish - Yaxshilangan versiya"""
        try:
            # Sana parametrini olish
            date_param = request.args.get('date')
            if date_param:
                target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            else:
                target_date = date.today()
            
            # Kunlik ma'lumotlarni olish
            daily_summary = face_system.data_manager.get_daily_summary(target_date)
            
            # Excel fayl yaratish
            wb = Workbook()
            ws = wb.active
            ws.title = f"Davomat_{target_date.strftime('%Y-%m-%d')}"
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Sarlavha
            ws.merge_cells('A1:F1')
            ws['A1'] = f"DAVOMAT HISOBOTI - {target_date.strftime('%d.%m.%Y')} ({daily_summary.get('day_name_uz', '')})"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws['A1'].alignment = center_alignment
            
            # Ma'lumot sarlavhasi
            ws.merge_cells('A2:F2')
            ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A2'].alignment = center_alignment
            ws['A2'].font = Font(size=10, italic=True)
            
            # Jadval sarlavhalari - Soddalashtirilgan
            headers = ['№', 'Foydalanuvchi', 'Birinchi Kirish', 'Oxirgi Chiqish', 'Jami Sessiyalar', 'Jami Vaqt']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Ma'lumotlarni yozish - Har bir user uchun FAQAT 1 ta qator
            row = 5
            user_number = 1
            
            if daily_summary.get('users'):
                for user_name, user_data in daily_summary['users'].items():
                    sessions = user_data.get('sessions', [])
                    total_minutes = user_data.get('total_minutes', 0)
                    
                    # Birinchi kirish va oxirgi chiqish
                    first_entry = sessions[0]['start'] if sessions else '-'
                    last_exit = sessions[-1]['end'] if sessions else '-'
                    
                    # Jami vaqtni soat:daqiqa formatida
                    hours = total_minutes // 60
                    minutes = total_minutes % 60
                    total_time_formatted = f"{hours}:{minutes:02d}"
                    
                    # Har bir foydalanuvchi uchun FAQAT 1 ta qator
                    ws.cell(row=row, column=1, value=user_number).border = border
                    ws.cell(row=row, column=2, value=user_name).border = border
                    ws.cell(row=row, column=3, value=first_entry).border = border
                    ws.cell(row=row, column=4, value=last_exit).border = border
                    ws.cell(row=row, column=5, value=len(sessions)).border = border
                    ws.cell(row=row, column=6, value=total_time_formatted).border = border
                    
                    # Alignment
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).alignment = center_alignment
                    
                    # Foydalanuvchi nomini bold qilish
                    ws.cell(row=row, column=2).font = Font(bold=True)
                    
                    row += 1
                    user_number += 1
            else:
                # Ma'lumot yo'q
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = "Bugun hech kim kirmagan"
                ws[f'A{row}'].alignment = center_alignment
                ws[f'A{row}'].font = Font(italic=True, color="999999")
            
            # Jami statistika
            if daily_summary.get('users'):
                total_users = len(daily_summary['users'])
                total_sessions = sum(len(user_data.get('sessions', [])) for user_data in daily_summary['users'].values())
                total_time = sum(user_data.get('total_minutes', 0) for user_data in daily_summary['users'].values())
                total_hours = total_time // 60
                total_mins = total_time % 60
                
                stats_row = row + 2
                
                # Statistika sarlavhasi
                ws.merge_cells(f'A{stats_row}:F{stats_row}')
                ws[f'A{stats_row}'] = "JAMI STATISTIKA"
                ws[f'A{stats_row}'].font = Font(bold=True, size=14, color="2F5597")
                ws[f'A{stats_row}'].fill = subheader_fill
                ws[f'A{stats_row}'].alignment = center_alignment
                ws[f'A{stats_row}'].border = border
                
                # Statistika ma'lumotlari
                stats_data = [
                    ("Jami foydalanuvchilar:", total_users),
                    ("Jami sessiyalar:", total_sessions),
                    ("Jami vaqt:", f"{total_hours}:{total_mins:02d}"),
                    ("O'rtacha vaqt/user:", f"{total_time//total_users if total_users > 0 else 0} daqiqa")
                ]
                
                for i, (label, value) in enumerate(stats_data):
                    stats_row += 1
                    ws[f'A{stats_row}'] = label
                    ws[f'B{stats_row}'] = value
                    ws[f'A{stats_row}'].font = Font(bold=True)
                    ws[f'A{stats_row}'].border = border
                    ws[f'B{stats_row}'].border = border
            
            # Ustun kengliklarini sozlash
            column_widths = [5, 25, 15, 15, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            # Faylni xotiraga saqlash
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Fayl nomini yaratish
            filename = f"Davomat_{target_date.strftime('%Y-%m-%d')}.xlsx"
            
            # Response yaratish
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return jsonify({'error': f'Excel export xatosi: {str(e)}'})
    
    return app